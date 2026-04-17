"""
Generate lead magnet files into public/downloads/.
Run from repo root: py -3 scripts/build_lead_magnets.py
Requires: openpyxl, fpdf2 (fpdf2 is auto-installed via pip if missing)
"""
from __future__ import annotations

import subprocess
import sys
import textwrap
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "public" / "downloads"

# Forge RPA site tokens (see src/styles/global.css @theme)
CHARCOAL = (26, 26, 46)
AMBER = (245, 158, 11)
AMBER_DARK = (217, 119, 6)

HIGH_ROI_GUIDE_MARKDOWN = textwrap.dedent(
    """
        # Five high-ROI finance processes to automate first

        **Forge RPA** — practical guide for finance and accounting leaders.  
        *Proven in finance; applicable to other high-volume operational processes.*

        ---

        ## 1. Accounts payable — intake, coding, and three-way match

        **Why ROI is high:** High transaction volume, structured data (PO, receipt, invoice), clear exception paths.  
        **Typical savings:** Dozens to hundreds of hours per month in larger AP teams when match and follow-up are manual.  
        **Implementation approach:** Stabilize matching rules and exception taxonomy; automate the “happy path”; queue exceptions for specialists.  
        **Caveats:** Vendor master quality and tax/VAT complexity can dominate effort — fix data before scaling bots.

        ---

        ## 2. Bank and GL reconciliations

        **Why ROI is high:** Recurring, deadline-driven, repetitive tie-outs across many accounts.  
        **Typical savings:** Material close compression and fewer late-night corrections.  
        **Implementation approach:** Standardize supporting workpapers; automate download, matching, and variance narratives where safe; escalate true breaks.  
        **Caveats:** One-off restructuring or new accounts need controlled onboarding into the automation library.

        ---

        ## 3. Month-end journal entry and consolidation support packs

        **Why ROI is high:** Same calendar-driven steps every period; heavy copy-paste between ERP and Excel.  
        **Typical savings:** Often 30–50%+ of manual pack time when reporting is template-driven.  
        **Implementation approach:** Treat packs as products — fixed inputs, fixed outputs, explicit validation checks; add controlled overrides for judgment lines.  
        **Caveats:** “Spreadsheet archaeology” should be replaced with governed models, not more hidden formulas.

        ---

        ## 4. Intercompany billings, allocations, and eliminations support

        **Why ROI is high:** Rules repeat across entities; errors are painful to unwind.  
        **Typical savings:** Fewer reversals and faster IC true-up during close.  
        **Implementation approach:** Document entity logic once; automate calculation and posting where policy is stable; keep audit trail explicit.  
        **Caveats:** Legal entity changes and tax-driven adjustments need governance gates.

        ---

        ## 5. Management and operational reporting from ERP subledgers

        **Why ROI is high:** Reporting pulls are constant; many are structured pivots and reconciliations to a single “source of truth.”  
        **Typical savings:** Frees FP&A and accounting from repetitive pulls into consistent dashboards.  
        **Implementation approach:** Lock definitions (metrics dictionary); automate extract-transform-load to reporting models; add reconciliation to GL.  
        **Caveats:** Bad mappings multiply — invest in data definitions before visual polish.

        ---

        ## How to use this list with the Assessment Framework

        1. Shortlist processes that match the patterns above in *your* environment.  
        2. Score each with the Excel framework (`forgerpa-finance-automation-assessment-framework.xlsx`).  
        3. Size hours, loaded cost, and adoption conservatively in the ROI calculator.  
        4. Book a discovery call on [forgerpa.com/book](https://forgerpa.com/book) if you want a second opinion on sequencing.

        ---

        *© Forge RPA. Provided for planning purposes. Not legal, tax, or investment advice.*
        """
).strip()


def ensure_fpdf() -> None:
    try:
        import fpdf  # noqa: F401
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "fpdf2", "-q"])


def style_header(cell, *, bold: bool = True, fill: str | None = "FFF4E6") -> None:
    cell.font = Font(bold=bold)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_assessment_workbook() -> None:
    wb = Workbook()
    # --- Instructions ---
    ws = wb.active
    ws.title = "Instructions"
    ws["A1"] = "Forge RPA — Finance Automation Assessment Framework"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A3"] = textwrap.dedent(
        """
        How to use this workbook
        1. List each candidate process on the \"Processes\" sheet (one row per process).
        2. Enter estimated manual hours per month and primary systems.
        3. Score each criterion from 1 (weak fit) to 5 (strong fit). Be honest about exceptions and judgment-heavy steps.
        4. The Weighted Score uses the default weights on the Processes sheet (you may change weights in row 2; keep them as positive decimals).
        5. Sort or filter by Weighted Score to build a defensible automation backlog.

        Criteria (columns D–I on Processes)
        • Volume & frequency — How often the work runs and how much labor it consumes.
        • Rule stability — How much is repeatable rules vs. case-by-case judgment.
        • Data & system accessibility — Can data be reached reliably (exports, APIs, controlled UI)?
        • Exception load — Volume and complexity of true exceptions.
        • Risk / control impact — Cost of error (misstatement, late filing, compliance).
        • Urgency / deadline pressure — Month-end, regulatory, or management time sensitivity.

        Disclaimer: Scores are directional planning aids, not guarantees of ROI or feasibility.
        """
    ).strip()
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 100

    # --- Processes ---
    wp = wb.create_sheet("Processes")
    headers = [
        "Process name",
        "Est. manual hours / month",
        "Primary systems / evidence",
        "Volume & frequency (1-5)",
        "Rule stability (1-5)",
        "Data & system access (1-5)",
        "Exception load (1-5) — higher score = fewer exceptions",
        "Risk / control impact (1-5)",
        "Urgency / deadlines (1-5)",
        "Weighted score",
        "Notes / next action",
    ]
    weights = [0.20, 0.25, 0.20, 0.10, 0.15, 0.10]
    for col, h in enumerate(headers, start=1):
        c = wp.cell(row=1, column=col, value=h)
        style_header(c)
    for i, w in enumerate(weights, start=4):
        wp.cell(row=2, column=i, value=w)
        wp.cell(row=2, column=i).number_format = "0.00"
    wp.cell(row=2, column=1, value="Default weights (edit D2:I2; positive numbers):")
    wp.cell(row=2, column=1).font = Font(bold=True)

    for col in range(1, 12):
        wp.column_dimensions[get_column_letter(col)].width = 22 if col <= 3 else 14
    wp.column_dimensions["A"].width = 28
    wp.column_dimensions["C"].width = 32
    wp.column_dimensions["K"].width = 36

    for r in range(5, 35):
        formula = (
            f'=IF(COUNT(D{r}:I{r})<6,"",SUMPRODUCT(D{r}:I{r},$D$2:$I$2)/SUM($D$2:$I$2))'
        )
        wp.cell(row=r, column=10, value=formula)

    # --- Interpretation ---
    wi = wb.create_sheet("Score guide")
    wi["A1"] = "Weighted score (approximate interpretation)"
    wi["A1"].font = Font(bold=True, size=14)
    rows = [
        ("4.3 – 5.0", "Strong candidate — prioritize for discovery / deeper sizing."),
        ("3.5 – 4.2", "Good candidate — validate exceptions, data, and ownership."),
        ("2.5 – 3.4", "Possible — often needs process cleanup or scope split first."),
        ("1.0 – 2.4", "Defer or redesign — high judgment, brittle data, or unclear ROI."),
    ]
    for i, (rng, txt) in enumerate(rows, start=3):
        wi.cell(row=i, column=1, value=rng)
        wi.cell(row=i, column=2, value=txt)
    wi.column_dimensions["A"].width = 14
    wi.column_dimensions["B"].width = 70

    OUT.mkdir(parents=True, exist_ok=True)
    path = OUT / "forgerpa-finance-automation-assessment-framework.xlsx"
    wb.save(path)
    print("Wrote", path)


def build_roi_workbook() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Calculator"

    ws["A1"] = "Forge RPA — Finance Automation ROI Calculator (illustrative)"
    ws["A1"].font = Font(size=15, bold=True)

    labels = [
        ("Hours reduced per month (after automation)", 4),
        ("Fully loaded hourly cost ($)", 5),
        ("Year-1 adoption (0–100%) — ramp", 6),
        ("One-time implementation cost ($)", 7),
        ("Annual run-state cost — licenses/support ($)", 8),
    ]
    for text, row in labels:
        ws.cell(row=row, column=1, value=text)
        ws.cell(row=row, column=2, value=0)

    ws["B4"], ws["B5"], ws["B6"], ws["B7"], ws["B8"] = 40, 85, 0.7, 25000, 4000
    for r in range(4, 9):
        ws.cell(row=r, column=2).number_format = "#,##0.00"

    ws["A10"] = "Monthly labor savings ($)"
    ws["B10"] = "=B4*B5*B6"
    ws["A11"] = "Annual labor savings ($)"
    ws["B11"] = "=B10*12"
    ws["A12"] = "Year-1 net benefit ($)"
    ws["B12"] = "=B11-B7-B8"
    ws["A13"] = "Simple payback on implementation (months)"
    ws["B13"] = "=IF(B10-B8/12<=0,\"n/a\",B7/(B10-B8/12))"

    for r in (10, 11, 12):
        ws.cell(row=r, column=2).number_format = "#,##0.00"
    ws["B13"].number_format = "0.0"

    ws["A15"] = textwrap.dedent(
        """
        Notes
        • Adjust adoption for partial-year rollout. This model does not discount cash flows.
        • Add implementation and run-state costs honestly; omit rows you do not use (set to 0).
        • Use for internal business-case discussion — not a promise of results.
        """
    ).strip()
    ws["A15"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 18

    wg = wb.create_sheet("Benchmarks (illustrative)")
    wg["A1"] = "Illustrative ranges only — anonymized patterns from finance automation work."
    wg["A1"].font = Font(italic=True)
    data = [
        ("Process family", "Typical manual hours / mo (wide range)", "Often automatable first"),
        ("AP invoice coding / 3-way match", "80–400+", "Data + rules heavy"),
        ("Bank / GL reconciliations", "40–200+", "High frequency, template-friendly"),
        ("Month-end JE packs / flux", "30–150+", "Often semi-structured"),
        ("Operational reporting packs", "25–120+", "Depends on data hygiene"),
    ]
    for r, row in enumerate(data, start=3):
        for c, val in enumerate(row, start=1):
            wg.cell(row=r, column=c, value=val)
    wg.column_dimensions["A"].width = 34
    wg.column_dimensions["B"].width = 38
    wg.column_dimensions["C"].width = 36

    path = OUT / "forgerpa-finance-automation-roi-calculator.xlsx"
    wb.save(path)
    print("Wrote", path)


def build_guide_markdown() -> None:
    path = OUT / "forgerpa-5-high-roi-finance-processes.md"
    path.write_text(HIGH_ROI_GUIDE_MARKDOWN + "\n", encoding="utf-8")
    print("Wrote", path)


def build_guide_pdf() -> None:
    ensure_fpdf()
    from fpdf import FPDF

    class GuidePDF(FPDF):
        def footer(self) -> None:
            self.set_y(-12)
            self.set_font("Helvetica", "", 8)
            self.set_text_color(100, 116, 139)
            self.cell(0, 8, f"Forge RPA  |  forgerpa.com/resources  |  Page {self.page_no()}", align="C")

    pdf = GuidePDF()
    pdf.set_auto_page_break(auto=True, margin=16)
    pdf.set_margins(18, 18, 18)
    pdf.add_page()
    ew = pdf.w - pdf.l_margin - pdf.r_margin

    # Header band (matches site charcoal + amber accent)
    pdf.set_fill_color(*CHARCOAL)
    pdf.rect(0, 0, 210, 36, "F")
    pdf.set_xy(pdf.l_margin, 9)
    pdf.set_font("Helvetica", "B", 11)
    pdf.set_text_color(*AMBER)
    pdf.cell(ew, 5, "Forge RPA", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "B", 15)
    pdf.set_text_color(255, 255, 255)
    pdf.multi_cell(ew, 6, "Five high-ROI finance processes to automate first")
    pdf.set_x(pdf.l_margin)
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(226, 232, 240)
    pdf.multi_cell(
        ew,
        4,
        "Practical guide for finance and accounting leaders. Proven in finance; applicable to other high-volume operational processes.",
    )

    pdf.set_y(42)
    pdf.set_x(pdf.l_margin)
    pdf.set_text_color(15, 23, 42)
    pdf.set_font("Helvetica", "", 10)

    def section(title: str, parts: list[tuple[str, str]]) -> None:
        pdf.ln(3)
        pdf.set_x(pdf.l_margin)
        pdf.set_font("Helvetica", "B", 12)
        pdf.set_text_color(*CHARCOAL)
        pdf.multi_cell(ew, 5, title)
        pdf.set_font("Helvetica", "", 10)
        pdf.set_text_color(30, 41, 59)
        for label, body in parts:
            pdf.set_x(pdf.l_margin)
            pdf.set_font("Helvetica", "B", 10)
            pdf.cell(ew, 5, label, new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("Helvetica", "", 10)
            pdf.multi_cell(ew, 4, body)
            pdf.ln(1)

    section(
        "1. Accounts payable - intake, coding, and three-way match",
        [
            ("Why ROI is high:", "High transaction volume, structured data (PO, receipt, invoice), clear exception paths."),
            ("Typical savings:", "Dozens to hundreds of hours per month in larger AP teams when match and follow-up are manual."),
            ("Implementation approach:", "Stabilize matching rules and exception taxonomy; automate the happy path; queue exceptions for specialists."),
            ("Caveats:", "Vendor master quality and tax/VAT complexity can dominate effort - fix data before scaling bots."),
        ],
    )
    section(
        "2. Bank and GL reconciliations",
        [
            ("Why ROI is high:", "Recurring, deadline-driven, repetitive tie-outs across many accounts."),
            ("Typical savings:", "Material close compression and fewer late-night corrections."),
            ("Implementation approach:", "Standardize supporting workpapers; automate download, matching, and variance narratives where safe; escalate true breaks."),
            ("Caveats:", "One-off restructuring or new accounts need controlled onboarding into the automation library."),
        ],
    )
    section(
        "3. Month-end journal entry and consolidation support packs",
        [
            ("Why ROI is high:", "Same calendar-driven steps every period; heavy copy-paste between ERP and Excel."),
            ("Typical savings:", "Often 30-50%+ of manual pack time when reporting is template-driven."),
            ("Implementation approach:", "Treat packs as products - fixed inputs, fixed outputs, explicit validation checks; add controlled overrides for judgment lines."),
            ("Caveats:", "Spreadsheet archaeology should be replaced with governed models, not more hidden formulas."),
        ],
    )
    section(
        "4. Intercompany billings, allocations, and eliminations support",
        [
            ("Why ROI is high:", "Rules repeat across entities; errors are painful to unwind."),
            ("Typical savings:", "Fewer reversals and faster IC true-up during close."),
            ("Implementation approach:", "Document entity logic once; automate calculation and posting where policy is stable; keep audit trail explicit."),
            ("Caveats:", "Legal entity changes and tax-driven adjustments need governance gates."),
        ],
    )
    section(
        "5. Management and operational reporting from ERP subledgers",
        [
            ("Why ROI is high:", "Reporting pulls are constant; many are structured pivots and reconciliations to a single source of truth."),
            ("Typical savings:", "Frees FP&A and accounting from repetitive pulls into consistent dashboards."),
            ("Implementation approach:", "Lock definitions (metrics dictionary); automate extract-transform-load to reporting models; add reconciliation to GL."),
            ("Caveats:", "Bad mappings multiply - invest in data definitions before visual polish."),
        ],
    )

    pdf.ln(2)
    pdf.set_x(pdf.l_margin)
    pdf.set_font("Helvetica", "B", 12)
    pdf.set_text_color(*CHARCOAL)
    pdf.cell(ew, 6, "How to use this list with the Assessment Framework", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(30, 41, 59)
    steps = [
        "Shortlist processes that match the patterns above in your environment.",
        "Score each with the Excel workbook forgerpa-finance-automation-assessment-framework.xlsx.",
        "Size hours, loaded cost, and adoption conservatively in forgerpa-finance-automation-roi-calculator.xlsx.",
        "Book a discovery call if you want a second opinion on sequencing.",
    ]
    for i, s in enumerate(steps, start=1):
        pdf.set_x(pdf.l_margin)
        pdf.multi_cell(ew, 4, f"{i}. {s}")
    pdf.ln(2)
    pdf.set_x(pdf.l_margin)
    pdf.set_text_color(*AMBER_DARK)
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(ew, 5, "Schedule a 30-minute discovery call:", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "U", 10)
    pdf.set_text_color(30, 64, 175)
    pdf.set_x(pdf.l_margin)
    pdf.cell(ew, 5, "https://forgerpa.com/book", new_x="LMARGIN", new_y="NEXT", link="https://forgerpa.com/book")

    pdf.ln(4)
    pdf.set_x(pdf.l_margin)
    pdf.set_font("Helvetica", "I", 8)
    pdf.set_text_color(100, 116, 139)
    pdf.multi_cell(ew, 4, "Provided for planning purposes. Not legal, tax, or investment advice.")

    path = OUT / "forgerpa-5-high-roi-finance-processes.pdf"
    pdf.output(str(path))
    print("Wrote", path)


def build_checklist_html() -> None:
    sections: list[tuple[str, list[str]]] = [
        (
            "Pre–month-end readiness (items 1–10)",
            [
                "Cutoff communications sent (AP, AR, procurement, treasury)",
                "Accrual templates prepared with clear ownership",
                "Preliminary FX rates captured where required",
                "Intercompany schedules updated for new activity",
                "Inventory / COGS cutoffs coordinated with ops",
                "Fixed assets rollforward reconciled to subledger",
                "Prepaids schedule current; amortization run reviewed",
                "Debt covenant reporting inputs identified early",
                "Payroll calendar aligned with GL periods",
                "Critical access and segregation of duties verified for close roles",
            ],
        ),
        (
            "Core close execution (items 11–22)",
            [
                "Cash position and bank cutoff confirmed",
                "AR aging and allowance model inputs frozen",
                "AP aging and duplicate invoice scan completed",
                "Revenue cutoffs tested (ship date / service period)",
                "Standard JE templates posted with review checklist",
                "Payroll and benefits interfaces tied out",
                "Inventory valuation and obsolescence review completed",
                "Project accounting WIP and percent-complete reviewed",
                "Equity and intercompany picks/eliminations prepared",
                "Tax provision inputs synchronized with GL",
                "Foreign subs consolidated with locked rates / policies",
                "Management review of unusual fluctuations documented",
            ],
        ),
        (
            "Reconciliations and subledgers (items 23–32)",
            [
                "Bank reconciliations completed with aged items cleared",
                "Credit card and petty cash reconciled",
                "AR subledger to GL tie-out",
                "AP subledger to GL tie-out",
                "Fixed assets rollforward ties to GL",
                "Prepaids and intangibles schedules tied",
                "Inventory perpetual to GL tie-out",
                "Payroll clearing accounts cleared",
                "Intercompany AR/AP nets and differences resolved",
                "Suspense / clearing accounts at or near zero with owners",
            ],
        ),
        (
            "Reporting, filings, and governance (items 33–40)",
            [
                "Financial statement tie-out (TB to statements)",
                "Footnote support packages complete",
                "Board / committee deck numbers reconciled to GL",
                "SEC / statutory filing numbers cross-walked to GL",
                "SOX controls evidence collected for material close controls",
                "Variance commentary drafted with Causes / actions",
                "Cash flow statement tie-out to bank and GL activity",
                "Commitments and contingencies disclosure checklist updated",
            ],
        ),
        (
            "Post-close and continuous improvement (items 41–47)",
            [
                "Close calendar retrospective — what slipped and why",
                "Automation backlog updated from manual pain points",
                "Process owners assigned for recurring exceptions",
                "Master data cleanup tickets created (vendors, customers, items)",
                "Training gaps logged for new hires or new systems",
                "Documentation updated for recurring JEs and close paths",
                "Lessons learned shared with FP&A and operations for next cycle",
            ],
        ),
    ]
    total = sum(len(pts) for _, pts in sections)
    assert total == 47, total

    sections_html = ""
    n = 0
    for title, pts in sections:
        chunk = "\n".join(
            f"<li><label><input type=\"checkbox\" /> {n + i + 1}. {text}</label></li>"
            for i, text in enumerate(pts)
        )
        n += len(pts)
        sections_html += f"<section class=\"block\"><h2>{title}</h2><ol class=\"chk\">{chunk}</ol></section>"

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <meta name="theme-color" content="#1a1a2e" />
  <title>Month-end close optimization checklist — Forge RPA</title>
  <link rel="preconnect" href="https://fonts.googleapis.com" />
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin />
  <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700;800&display=swap" rel="stylesheet" />
  <style>
    :root {{
      --charcoal: #1a1a2e;
      --charcoal-light: #2d2d44;
      --amber: #f59e0b;
      --amber-dark: #d97706;
      --amber-tint: #fffbeb;
      --slate-700: #334155;
      --slate-500: #64748b;
      font-family: Inter, system-ui, sans-serif;
      color: var(--slate-700);
    }}
    body {{ max-width: 48rem; margin: 2rem auto; padding: 0 1.25rem 3rem; background: #fff; }}
    h1 {{
      font-size: 1.75rem;
      font-weight: 800;
      color: var(--charcoal);
      margin-bottom: 0.35rem;
      padding-bottom: 0.4rem;
      border-bottom: 4px solid var(--amber);
      display: inline-block;
    }}
    .sub {{ color: var(--slate-500); margin-bottom: 1rem; max-width: 40rem; line-height: 1.45; }}
    .sub .brand {{ color: var(--amber-dark); font-weight: 700; }}
    .meta-bar {{
      display: grid;
      grid-template-columns: 1fr;
      gap: 0.75rem;
      margin: 1rem 0 1.25rem;
      padding: 1rem 1.1rem;
      background: #f8fafc;
      border: 1px solid #e2e8f0;
      border-radius: 0.65rem;
    }}
    @media (min-width: 640px) {{
      .meta-bar {{ grid-template-columns: 1fr 1fr 1fr; }}
    }}
    .meta-field {{ font-size: 0.8rem; color: var(--slate-500); }}
    .meta-label {{ display: block; font-weight: 600; color: var(--charcoal); margin-bottom: 0.2rem; }}
    .meta-line {{ display: block; min-height: 1.35rem; border-bottom: 1px solid #cbd5e1; }}
    .hint {{
      background: var(--amber-tint);
      border: 1px solid var(--amber);
      padding: 0.85rem 1rem;
      border-radius: 0.65rem;
      margin-bottom: 1.25rem;
      font-size: 0.9rem;
      line-height: 1.45;
      color: var(--slate-700);
    }}
    .cta-strip {{
      background: linear-gradient(135deg, var(--charcoal) 0%, var(--charcoal-light) 100%);
      color: #fff;
      padding: 1.25rem 1.35rem;
      border-radius: 0.85rem;
      margin: 0 0 1.75rem;
      border: 2px solid var(--amber);
      box-shadow: 0 10px 25px rgba(26, 26, 46, 0.12);
    }}
    .cta-strip h2 {{ margin: 0 0 0.35rem; font-size: 1.15rem; font-weight: 700; color: #fff; border: none; padding: 0; }}
    .cta-strip p {{ margin: 0.35rem 0 0; font-size: 0.95rem; color: #e2e8f0; line-height: 1.45; }}
    .cta-strip .btn-book {{
      display: inline-block;
      margin-top: 1rem;
      padding: 0.7rem 1.35rem;
      background: var(--amber);
      color: var(--charcoal);
      font-weight: 800;
      font-size: 0.95rem;
      border-radius: 0.5rem;
      text-decoration: none;
      border: 2px solid var(--amber-dark);
    }}
    .cta-strip .btn-book:hover {{ background: var(--amber-dark); color: #fff; }}
    .print-only-url {{ display: none; font-size: 0.75rem; color: #cbd5e1; margin-top: 0.5rem; }}
    section.block {{ margin-bottom: 2rem; page-break-inside: avoid; }}
    h2 {{
      font-size: 1.05rem;
      font-weight: 700;
      margin-bottom: 0.65rem;
      color: var(--charcoal);
      border-left: 4px solid var(--amber);
      padding-left: 0.55rem;
    }}
    ol.chk {{ list-style: none; padding: 0; margin: 0; }}
    ol.chk li {{ margin: 0.45rem 0; line-height: 1.35; color: var(--slate-700); }}
    label {{ cursor: pointer; }}
    footer {{ margin-top: 2.5rem; font-size: 0.8rem; color: #94a3b8; border-top: 1px solid #e2e8f0; padding-top: 1rem; }}
    footer a {{ color: var(--amber-dark); font-weight: 600; text-decoration: none; }}
    footer a:hover {{ text-decoration: underline; }}
    @media print {{
      body {{ margin: 0.5rem; }}
      .hint {{ border-color: #ccc; background: #fafafa; }}
      .cta-strip {{ box-shadow: none; break-inside: avoid; }}
      .cta-strip .btn-book {{ border: 1px solid #000; color: #000; background: #fbbf24; }}
      .print-only-url {{ display: block; color: #333; }}
    }}
  </style>
</head>
<body>
  <h1>Month-end close optimization checklist</h1>
  <p class="sub"><span class="brand">Forge RPA</span> &mdash; 47 checkpoints with automation notes. Use <strong>Print &rarr; Save as PDF</strong> in your browser to keep a dated copy for your files.</p>
  <div class="meta-bar">
    <div class="meta-field"><span class="meta-label">Entity / BU (optional)</span><span class="meta-line"></span></div>
    <div class="meta-field"><span class="meta-label">Close period</span><span class="meta-line"></span></div>
    <div class="meta-field"><span class="meta-label">Prepared by</span><span class="meta-line"></span></div>
  </div>
  <p class="hint"><strong>Automation angle:</strong> mark items that are mostly rules plus repeatable data pulls &mdash; those are usually the best early automation candidates. Judgment-heavy items may still benefit from workflow and evidence packaging.</p>
  <div class="cta-strip">
    <h2>Want help turning checks into a roadmap?</h2>
    <p>Book a free 30-minute discovery call. We will help you prioritize what to automate first and what to stabilize before you invest.</p>
    <a class="btn-book" href="https://forgerpa.com/book">Book a discovery call</a>
    <p class="print-only-url">Booking link: https://forgerpa.com/book</p>
  </div>
  {sections_html}
  <footer>&copy; Forge RPA. For planning use. More resources: <a href="https://forgerpa.com/resources">forgerpa.com/resources</a> &middot; <a href="https://forgerpa.com/book">forgerpa.com/book</a></footer>
</body>
</html>"""
    path = OUT / "forgerpa-month-end-close-checklist.html"
    path.write_text(html, encoding="utf-8")
    print("Wrote", path)


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    build_assessment_workbook()
    build_roi_workbook()
    build_guide_markdown()
    build_guide_pdf()
    build_checklist_html()


if __name__ == "__main__":
    main()
